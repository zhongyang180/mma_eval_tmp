import json
import base64
import warnings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor, AnchorMarker
from typing import Dict, Any, List, Tuple
from datetime import datetime
import os
import time

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

def excel_to_json(excel_path: str) -> Tuple[Dict[str, Any], List[Dict[str, str]]]:
    """
    将Excel文件转换为符合指定格式的JSON，并正确处理图片
    返回格式: (
        {
            "title": "文件名",
            "content": [
                {"type": "text", "content": "文本行1"},
                {"type": "image", "name": "#/pictures/0", "description": "[待生成]"},
                {"type": "table", "content": {表格数据}},
                # 更多内容元素...
            ]
        },
        [
            {
                "type": "image",
                "name": "#/pictures/0",
                "uri": "data:image/png;base64,...",
                "description": "[待生成]"
            }
        ]
    )
    """
    # 获取文件名作为标题（不带扩展名）
    title = os.path.splitext(os.path.basename(excel_path))[0]
    
    # 初始化内容列表和图片列表
    content = []
    all_images = []
    # 全局图片索引
    image_index = 0

    # 加载工作簿
    workbook = load_workbook(
        excel_path, 
        data_only=True,
        keep_vba=False  # 禁止加载VBA内容
    )

    # 处理每个sheet
    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]

        # 添加sheet标题
        content.append({
            "type": "text",
            "content": f"工作表 {sheet_index + 1}: {sheet_name}"
        })

        # 收集当前sheet的所有图片
        sheet_images = []
        # 使用 sheet._images 获取图片（这是正确的方法）
        if hasattr(sheet, '_images') and sheet._images:
            for img_obj in sheet._images:
                # 确保是图片对象
                if not isinstance(img_obj, Image):
                    continue
                    
                # 获取图片位置（左上角单元格）
                anchor = img_obj.anchor
                row = 1
                col = 1
                
                # 处理不同类型的锚点
                if isinstance(anchor, OneCellAnchor):
                    # 单单元格锚点
                    if hasattr(anchor, 'top_left'):
                        cell = anchor.top_left
                    else:
                        cell = anchor._from
                    row = cell.row
                    col = cell.col
                elif isinstance(anchor, TwoCellAnchor):
                    # 双单元格锚点
                    if hasattr(anchor, 'top_left'):
                        cell = anchor.top_left
                    else:
                        cell = anchor._from
                    row = cell.row
                    col = cell.col
                elif isinstance(anchor, AnchorMarker):
                    # 锚点标记
                    row = anchor.row
                    col = anchor.col
                
                sheet_images.append((row, col, img_obj))
        
        # 按行和列排序图片（从上到下，从左到右）
        sheet_images.sort(key=lambda x: (x[0], x[1]))
        
        # 处理当前sheet的图片
        for row, col, img_obj in sheet_images:
            # 生成图片的base64编码
            try:
                # 获取图片数据
                img_data = img_obj._data()
                
                # 获取图片格式
                img_format = 'png'  # 默认格式
                if hasattr(img_obj, 'format') and img_obj.format:
                    img_format = img_obj.format.lower()
                    # 处理JPEG格式的别名
                    if img_format == 'jpeg':
                        img_format = 'jpg'
                
                # 生成Base64编码的图片URI
                base64_encoded = base64.b64encode(img_data).decode('utf-8')
                uri = f"data:image/{img_format};base64,{base64_encoded}"
                
                # 添加到图片列表
                image_info = {
                    "type": "image",
                    "name": f"#/pictures/{image_index}",
                    "uri": uri,
                    "description": "[待生成]"
                }
                all_images.append(image_info)
                
                # 在内容中添加图片引用
                content.append({
                    "type": "image",
                    "name": f"#/pictures/{image_index}",
                    "description": "[待生成]"
                })
                
                image_index += 1
            except Exception as e:
                print(f"处理图片失败: {str(e)}")
                continue

        # 获取合并单元格信息
        merged_ranges = {}
        if sheet.merged_cells:
            for merged_range in sheet.merged_cells.ranges:
                min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
                merged_value = sheet.cell(row=min_row, column=min_col).value
                merged_ranges[(min_row, min_col, max_row, max_col)] = merged_value

        # 创建合并单元格映射
        merged_map = {}
        for (min_row, min_col, max_row, max_col), value in merged_ranges.items():
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    merged_map[(row_idx, col_idx)] = value

        # 计算表格维度
        max_row = sheet.max_row
        max_col = sheet.max_column

        # 提取表格所有数据
        all_rows = []
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                # 检查是否是合并单元格
                if (row_idx, col_idx) in merged_map:
                    cell_value = merged_map[(row_idx, col_idx)]
                else:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value

                # 处理空值
                if cell_value is None:
                    cell_value = ""

                # 保留原始数据类型
                row_data.append(cell_value)
            all_rows.append(row_data)

        # 构建表格内容
        table_content = {
            "dimensions": {
                "rows": len(all_rows),
                "columns": max_col
            },
            "headers": all_rows[0] if all_rows else [],
            "data": all_rows[1:] if len(all_rows) > 1 else []
        }

        # 添加到内容列表
        content.append({
            "type": "table",
            "content": table_content
        })

    # 添加结束文本
    content.append({
        "type": "text",
        "content": f"文件 '{title}' 中共有 {len(workbook.sheetnames)} 个工作表"
    })

    return {
        "title": title,
        "content": content
    }, all_images

def save_json(data: Any, json_path: str) -> None:
    """保存JSON数据到文件"""
    # 确保目录存在
    os.makedirs(os.path.dirname(json_path), exist_ok=True)

    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def analyze_excel(excel_path: str, json_path: str ="tmp/chunk.json") -> Dict[str, Any]:
    """完整的Excel分析流程"""
    print("\n" + "="*50)
    print("开始Excel分析流程")
    start_time = time.time()
    start_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"开始时间: {start_datetime}")
    print(f"输入文件: {excel_path}")
    print(f"输出文件: {json_path}")

    try:
        # 转换Excel到JSON格式
        json_data, images_list = excel_to_json(excel_path)

        # 保存主JSON
        save_json(json_data, json_path)
        
        # 保存图片JSON
        base_name = os.path.splitext(json_path)[0]
        images_json_path = f"{base_name}_images.json"
        save_json(images_list, images_json_path)

        end_time = time.time()
        end_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        duration = end_time - start_time

        print(f"结束时间: {end_datetime}")
        print(f"耗时: {duration:.2f} 秒")
        print(f"发现图片数量: {len(images_list)}")

        # 计算表格行数
        total_rows = 0
        for item in json_data["content"]:
            if item["type"] == "table":
                total_rows += item["content"]["dimensions"]["rows"]

        print(f"转换完成，共处理 {total_rows} 行数据")
        print(f"主JSON已保存到: {json_path}")
        print(f"图片JSON已保存到: {images_json_path}")

        return json_data

    except Exception as e:
        print(f"Excel分析失败: {str(e)}")
        raise

if __name__ == "__main__":
    # 设置文件路径
    excel_file = "/data5/sby/mma_eval/tmp/test.xlsx"
    x = extract_cell_triplets(excel_file)

    print(x)